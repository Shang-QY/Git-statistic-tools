import git
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 配置
repo_path = 'Intel'  # 你的Git仓库路径
input_file = 'cherry-pick-progress.txt'  # 包含提交哈希的文本文件
output_file = 'commits-new5.xlsx'  # 输出的Excel文件

# 初始化Git仓库
repo = git.Repo(repo_path)

# 创建一个新的Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "Commits"

# 设置Excel表头
headers = [
    'Commit Hash', 'Commit Message Title', 'Detailed Commit Message', 'Modified Files',
    'Lines Added', 'Lines Deleted',
    'AlderLakeBoardPkg Added', 'AlderLakeBoardPkg Modified',
    'AlderLakePlatSamplePkg Added', 'AlderLakePlatSamplePkg Modified',
    'AlderLakeFspPkg Added', 'AlderLakeFspPkg Modified',
    'ClientOneSiliconPkg Added', 'ClientOneSiliconPkg Modified'
]
ws.append(headers)

# 读取提交哈希列表
with open(input_file, 'r') as file:
    commit_hashes = [line.strip() for line in file if line.strip()]

# 定义一个函数来统计特定路径下的增加和修改行数
def count_lines_in_path(stats, path_prefix):
    added = sum(stats['insertions'] for file_path, stats in stats.items() if file_path.startswith(path_prefix))
    modified = sum(stats['deletions'] for file_path, stats in stats.items() if file_path.startswith(path_prefix))
    return added, modified

# 查询每个提交的详细信息
for commit_hash in commit_hashes:
    print(commit_hash)
    try:
        commit = repo.commit(commit_hash)
        commit_message_title = commit.message.strip().split('\n')[0]  # 获取提交消息的第一行
        detailed_commit_message = commit.message.strip()  # 获取完整的提交消息
        # 获取修改的文件列表及其统计数据
        modified_files_stats = []
        total_lines_added = 0
        total_lines_deleted = 0
        for file_path, stats in commit.stats.files.items():
            modified_files_stats.append(f"{file_path} (+{stats['insertions']}/-{stats['deletions']})")
            total_lines_added += stats['insertions']
            total_lines_deleted += stats['deletions']
        modified_files = '\n'.join(modified_files_stats)  # 每个文件后换行
        # 统计特定路径下的增加和修改行数
        al_board_pkg_added, al_board_pkg_modified = count_lines_in_path(commit.stats.files, 'AlderLakeBoardPkg/')
        al_plat_pkg_added, al_plat_pkg_modified = count_lines_in_path(commit.stats.files, 'AlderLakePlatSamplePkg/')
        al_fsp_pkg_added, al_fsp_pkg_modified = count_lines_in_path(commit.stats.files, 'AlderLakeFspPkg/')
        client_one_pkg_added, client_one_pkg_modified = count_lines_in_path(commit.stats.files, 'ClientOneSiliconPkg/')
        # 创建一个包含提交信息的行
        row = [
            commit_hash, commit_message_title, detailed_commit_message, modified_files,
            total_lines_added, total_lines_deleted,
            al_board_pkg_added, al_board_pkg_modified,
            al_plat_pkg_added, al_plat_pkg_modified,
            al_fsp_pkg_added, al_fsp_pkg_modified,
            client_one_pkg_added, client_one_pkg_modified
        ]
        ws.append(row)
        # 获取当前行号
        current_row = ws.max_row
        # 设置单元格的自动换行
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True)
    except Exception as e:
        print(f"Error processing commit {commit_hash}: {e}")

# 调整列宽
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# 保存Excel文件
wb.save(filename=output_file)
print(f"Excel file '{output_file}' has been created with commit details.")
